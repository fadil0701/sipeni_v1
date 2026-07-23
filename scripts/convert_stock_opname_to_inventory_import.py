#!/usr/bin/env python3
"""
Konversi Stock Opname ATK/Cetakan/ART → baris import Data Inventory (PERSEDIAAN),
lalu digabung ke file import ASET (mis. import_inventory_aset_kib_b_prod.xlsx).

Contoh:
  py scripts/convert_stock_opname_to_inventory_import.py

  py scripts/convert_stock_opname_to_inventory_import.py ^
      --stock \"database/seeders/data/Stock Opname ATK CETAKAN ART 2026 (5).xlsx\" ^
      --target \"database/seeders/data/import_inventory_aset_kib_b_prod.xlsx\" ^
      --id-gudang 24
"""

from __future__ import annotations

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

# Reuse normalisasi satuan & klasifikasi dari enrich Kemendagri
ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))
from enrich_kemendagri_persediaan_farmasi import (  # noqa: E402
    SECTION_ALIASES,
    JUNK_NAMA,
    STOCK_SHEET,
    normalize_satuan,
)

DEFAULT_STOCK = ROOT / "database/seeders/data/Stock Opname ATK CETAKAN ART 2026 (5).xlsx"
DEFAULT_TARGET = ROOT / "database/seeders/data/import_inventory_aset_kib_b_prod.xlsx"
DEFAULT_KEMENDAGRI = ROOT / "database/seeders/data/kemendagri_import_sheet6_objek_filtered_v2.xlsx"

HEADERS = [
    "id_data_barang",
    "id_gudang",
    "id_anggaran",
    "id_sub_kegiatan",
    "jenis_inventory",
    "jenis_barang",
    "tahun_anggaran",
    "qty_input",
    "id_satuan",
    "harga_satuan",
    "merk",
    "tipe",
    "spesifikasi",
    "tahun_produksi",
    "nama_penyedia",
    "no_seri",
    "no_batch",
    "tanggal_kedaluwarsa",
    "status_inventory",
    "upload_foto",
    "upload_dokumen",
]

# Urutan sama MasterSatuanSeeder (id lokal/production setelah seed bersih).
SATUAN_IDS = {
    "Unit": 1,
    "PCS": 2,
    "Buah": 3,
    "Pasang": 4,
    "Set": 5,
    "Lusin": 6,
    "Kodi": 7,
    "Gross": 8,
    "Rim": 9,
    "Lembar": 10,
    "Roll": 11,
    "Ikat": 12,
    "Batang": 13,
    "Potong": 14,
    "Ekor": 15,
    "Biji": 16,
    "Helai": 17,
    "Box": 18,
    "Dus": 19,
    "Karton": 20,
    "Pak": 21,
    "Bundle": 22,
    "Pallet": 23,
    "Karung": 24,
    "Kaleng": 25,
    "Tabung": 26,
    "Jerigen": 27,
    "Drum": 28,
    "Bag": 29,
    "Strip": 30,
    "Blister": 31,
    "Sachet": 32,
    "Pouch": 33,
    "Tube": 34,
    "Botol": 35,
    "Ampul": 36,
    "Vial": 37,
    "Flakon": 38,
    "Prefilled Syringe": 39,
    "Tablet": 40,
    "Kaplet": 41,
    "Kapsul": 42,
    "Suppositoria": 43,
    "Ovula": 44,
    "Sirup": 45,
    "Eliksir": 46,
    "Tetes": 47,
    "Drop": 48,
    "Puff": 49,
    "Sachet (bubuk)": 50,
    "Vial (serbuk)": 51,
    "kg": 52,
    "g": 53,
    "mg": 54,
    "mcg": 55,
    "Mikrogram": 56,
    "Ton": 57,
    "L": 58,
    "mL": 59,
    "cc": 60,
    "m": 61,
    "cm": 62,
    "mm": 63,
    "km": 64,
    "m²": 65,
    "Dosis": 66,
    "IU": 67,
    "Test": 68,
    "Kit": 69,
    "Slide": 70,
    "Panel": 71,
    "Rack": 72,
    "Tray": 73,
    "Pak isi": 74,
    "Book": 75,
    "Pad": 76,
}


def resolve_path(path: Path) -> Path:
    return path if path.is_absolute() else (ROOT / path)


def norm_space(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def year_from(value) -> int | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.year
    text = norm_space(value)
    m = re.search(r"(19|20)\d{2}", text)
    return int(m.group(0)) if m else None


def parse_money(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = norm_space(value).replace("Rp", "").replace("rp", "").replace(" ", "")
    # 24.500,00 (ID) vs 24500.00
    if re.search(r",\d{2}$", text) and "." in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", "")
    try:
        return float(text)
    except ValueError:
        return 0.0


def parse_qty(value) -> float:
    if value is None or value == "" or value == "-":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = norm_space(value).replace("-", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def jenis_barang_from_section(section_label: str) -> str:
    """Map header SO → nilai allowed importer PERSEDIAAN."""
    u = section_label.upper()
    if "KHUSUS" in u:
        return "CETAK KHUSUS"  # ejaan di InventoryDataImportController
    if "CETAKAN" in u or u == "CETAKAN":
        return "CETAKAN UMUM"
    if "ART" in u or "RUMAH TANGGA" in u:
        return "ART"
    return "ATK"


def anggaran_from_source(text: str, id_apbd: int, id_blud: int, id_hibah: int) -> int:
    u = text.upper()
    if "HIBAH" in u:
        return id_hibah
    if "BLUD" in u:
        return id_blud
    return id_apbd


def load_nama_to_kode(kemendagri: Path) -> dict[str, str]:
    wb = load_workbook(kemendagri, read_only=True, data_only=True)
    if "data_barang" not in wb.sheetnames:
        wb.close()
        raise RuntimeError(f"Sheet data_barang tidak ada di {kemendagri}")
    ws = wb["data_barang"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    mapping: dict[str, str] = {}
    for row in rows[1:]:
        if not row or not row[0] or not row[1]:
            continue
        kode = str(row[0]).strip()
        nama = norm_space(row[1]).upper()
        if nama and nama not in mapping:
            mapping[nama] = kode
    return mapping


def parse_stock_lots(stock_path: Path) -> list[dict]:
    """
    Parse SO sheet menjadi lot persediaan:
    nama, satuan, harga, tahun, qty (saldo awal), sumber, jenis_barang.
    Baris lanjutan (nama kosong) mewarisi nama sebelumnya.
    """
    wb = load_workbook(stock_path, data_only=True)
    if STOCK_SHEET not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{STOCK_SHEET}' tidak ditemukan")
    ws = wb[STOCK_SHEET]

    current_section: str | None = None
    last_nama: str | None = None
    lots: list[dict] = []

    for row in ws.iter_rows(min_row=1, max_col=8, values_only=True):
        no, _kode, nama, satuan, harga, tahun_src, qty, _total = (list(row) + [None] * 8)[:8]
        nama_s = norm_space(nama)
        satuan_s = norm_space(satuan)
        no_s = norm_space(no)
        src_s = norm_space(tahun_src)

        if nama_s:
            upper = nama_s.upper()
            if upper in SECTION_ALIASES or any(upper.startswith(k) for k in SECTION_ALIASES):
                if (
                    re.fullmatch(r"[IVX]+", no_s)
                    or no_s in {"", "."}
                    or upper in SECTION_ALIASES
                ):
                    for alias in SECTION_ALIASES:
                        if upper == alias or upper.startswith(alias):
                            current_section = alias
                            last_nama = None
                            break
                    continue

        if current_section is None:
            continue

        if nama_s.upper().startswith("JUMLAH TOTAL"):
            current_section = None
            last_nama = None
            continue

        if not satuan_s:
            continue

        if nama_s:
            if nama_s.upper() in JUNK_NAMA:
                continue
            if re.fullmatch(r"[\d.,]+", nama_s):
                continue
            last_nama = nama_s
        elif not last_nama:
            continue

        qty_v = parse_qty(qty)
        if qty_v <= 0:
            continue

        harga_v = parse_money(harga)
        tahun = year_from(tahun_src) or 2026
        jenis = jenis_barang_from_section(current_section)

        lots.append(
            {
                "nama": last_nama,
                "satuan": normalize_satuan(satuan_s),
                "harga": round(harga_v, 2),
                "tahun": tahun,
                "qty": qty_v,
                "sumber": src_s,
                "jenis_barang": jenis,
            }
        )

    wb.close()
    return lots


def aggregate_lots(lots: list[dict]) -> list[dict]:
    """Gabung lot sama nama+harga+tahun+satuan+jenis+sumber_anggaran_key."""
    grouped: dict[tuple, dict] = {}
    order: list[tuple] = []
    for lot in lots:
        # sumber hanya untuk deteksi anggaran; agregasi tanpa teks sumber penuh
        key = (
            lot["nama"].upper(),
            lot["satuan"],
            lot["harga"],
            lot["tahun"],
            lot["jenis_barang"],
            "BLUD" if "BLUD" in lot["sumber"].upper() else (
                "HIBAH" if "HIBAH" in lot["sumber"].upper() else "APBD"
            ),
        )
        if key not in grouped:
            grouped[key] = dict(lot)
            order.append(key)
        else:
            grouped[key]["qty"] += lot["qty"]
    return [grouped[k] for k in order]


def build_rows(
    lots: list[dict],
    nama_to_kode: dict[str, str],
    *,
    id_gudang: int,
    id_apbd: int,
    id_blud: int,
    id_hibah: int,
    tahun_anggaran_default: int,
) -> tuple[list[list], list[str], list[str]]:
    rows: list[list] = []
    missing: list[str] = []
    skipped_satuan: list[str] = []
    seen_missing: set[str] = set()

    for lot in lots:
        nama_key = lot["nama"].upper()
        kode = nama_to_kode.get(nama_key)
        if not kode:
            if nama_key not in seen_missing:
                missing.append(lot["nama"])
                seen_missing.add(nama_key)
            continue

        satuan_name = lot["satuan"]
        id_satuan = SATUAN_IDS.get(satuan_name)
        if not id_satuan:
            skipped_satuan.append(f"{lot['nama']} ({satuan_name})")
            id_satuan = 1  # fallback Unit

        tahun_hist = int(lot["tahun"])
        tahun_anggaran = tahun_hist
        tahun_produksi = tahun_hist
        if tahun_anggaran < 2000 or tahun_anggaran > 2100:
            tahun_anggaran = tahun_anggaran_default

        id_anggaran = anggaran_from_source(lot["sumber"], id_apbd, id_blud, id_hibah)

        rows.append(
            [
                kode,
                id_gudang,
                id_anggaran,
                None,
                "PERSEDIAAN",
                lot["jenis_barang"],
                tahun_anggaran,
                int(lot["qty"]) if float(lot["qty"]).is_integer() else lot["qty"],
                id_satuan,
                lot["harga"],
                None,
                None,
                lot["nama"],
                tahun_produksi,
                lot["sumber"] or None,
                None,
                None,
                None,
                "AKTIF",
                None,
                None,
            ]
        )

    return rows, missing, skipped_satuan


def merge_into_target(target: Path, new_rows: list[list], meta: dict) -> None:
    if target.exists():
        wb = load_workbook(target)
        ws = wb.active
        # hapus baris PERSEDIAAN lama (re-run aman)
        header = [str(c.value or "").lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        try:
            j_idx = header.index("jenis_inventory")
        except ValueError:
            j_idx = 4
        to_delete = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row and str(row[j_idx] or "").upper() == "PERSEDIAAN":
                to_delete.append(idx)
        for idx in reversed(to_delete):
            ws.delete_rows(idx, 1)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "data_inventory"
        ws.append(HEADERS)

    for row in new_rows:
        ws.append(row)

    # petunjuk
    if "petunjuk" in wb.sheetnames:
        tip = wb["petunjuk"]
        tip.append([])
        tip.append([f"PERSEDIAAN ditambah: {meta['persediaan_rows']} baris dari Stock Opname"])
        tip.append([f"Sumber SO: {meta['stock']}"])
        tip.append([f"id_gudang PERSEDIAAN={meta['id_gudang']} (production: GUDANG PERSEDIAAN)"])
        tip.append([f"Dibuat: {meta['created_at']}"])
    else:
        tip = wb.create_sheet("petunjuk")
        tip.append([f"PERSEDIAAN: {meta['persediaan_rows']} baris | {meta['created_at']}"])

    target.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target)
    wb.close()


def main() -> int:
    parser = argparse.ArgumentParser(description="Stock Opname → import inventory PERSEDIAAN")
    parser.add_argument("--stock", type=Path, default=DEFAULT_STOCK)
    parser.add_argument("--target", type=Path, default=DEFAULT_TARGET)
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Path output (default: sama dengan --target). Pakai jika target terkunci Excel.",
    )
    parser.add_argument("--kemendagri", type=Path, default=DEFAULT_KEMENDAGRI)
    parser.add_argument("--id-gudang", type=int, default=24, help="Gudang PUSAT PERSEDIAAN (prod=24)")
    parser.add_argument("--id-apbd", type=int, default=1)
    parser.add_argument("--id-blud", type=int, default=3)
    parser.add_argument("--id-hibah", type=int, default=4)
    parser.add_argument("--tahun-anggaran", type=int, default=datetime.now().year)
    args = parser.parse_args()

    stock = resolve_path(args.stock)
    target = resolve_path(args.target)
    output = resolve_path(args.output) if args.output else target
    kemendagri = resolve_path(args.kemendagri)

    for path, label in [(stock, "Stock Opname"), (kemendagri, "Kemendagri")]:
        if not path.exists():
            print(f"{label} tidak ditemukan: {path}", file=sys.stderr)
            return 1

    if not target.exists():
        print(f"Target tidak ditemukan: {target}", file=sys.stderr)
        return 1

    nama_to_kode = load_nama_to_kode(kemendagri)
    lots = parse_stock_lots(stock)
    lots = aggregate_lots(lots)
    rows, missing, skipped_satuan = build_rows(
        lots,
        nama_to_kode,
        id_gudang=args.id_gudang,
        id_apbd=args.id_apbd,
        id_blud=args.id_blud,
        id_hibah=args.id_hibah,
        tahun_anggaran_default=args.tahun_anggaran,
    )

    meta = {
        "stock": str(stock),
        "id_gudang": args.id_gudang,
        "persediaan_rows": len(rows),
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }

    if output.resolve() != target.resolve():
        from shutil import copy2

        copy2(target, output)

    merge_into_target(output, rows, meta)

    print("Selesai.")
    print(f"  Stock lots (agregat): {len(lots)}")
    print(f"  PERSEDIAAN rows    : {len(rows)}")
    print(f"  Output             : {output}")
    if missing:
        print(f"  Nama tanpa master  : {len(missing)} (lewati)")
        for n in missing[:15]:
            print(f"    - {n}")
        if len(missing) > 15:
            print(f"    ... +{len(missing) - 15} lagi")
    if skipped_satuan:
        print(f"  Satuan fallback Unit: {len(skipped_satuan)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
