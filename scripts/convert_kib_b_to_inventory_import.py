#!/usr/bin/env python3
"""
Konversi file KIB B (BMD) → template import Data Inventory (jenis ASET).

Output mengikuti header resmi sheet `data_inventory`:
  database/seeders/data/template_import_inventory_data.xlsx

Baris digabung (default) berdasarkan:
  KOBAR + merk + tahun pembelian + harga_satuan
Jika harga berbeda (meski merk & tahun sama), tetap baris terpisah.
Gunakan --no-aggregate untuk 1 baris KIB = 1 baris import.

Contoh pemakaian:
  py scripts/convert_kib_b_to_inventory_import.py

  py scripts/convert_kib_b_to_inventory_import.py ^
      --source "database/seeders/data/KIB B.xlsx" ^
      --output "database/seeders/data/import_inventory_aset_kib_b.xlsx"

  py scripts/convert_kib_b_to_inventory_import.py ^
      --aggregate-import "database/seeders/data/import_inventory_aset_kib_b_fixed.xlsx" ^
      --output "database/seeders/data/import_inventory_aset_kib_b_fixed.xlsx"

Ketergantungan: pip install openpyxl
"""

from __future__ import annotations

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("Modul openpyxl belum terpasang. Jalankan: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = ROOT / "database" / "seeders" / "data" / "KIB B.xlsx"
DEFAULT_OUTPUT = ROOT / "database" / "seeders" / "data" / "import_inventory_aset_kib_b.xlsx"

HEADERS = [
    "id_data_barang",
    "id_gudang",
    "id_anggaran",
    "id_sub_kegiatan",
    "jenis_inventory",
    "jenis_barang",
    "tahun_anggaran",
    "qty_input",
    "id_satuan",
    "harga_satuan",
    "merk",
    "tipe",
    "spesifikasi",
    "tahun_produksi",
    "nama_penyedia",
    "no_seri",
    "no_batch",
    "tanggal_kedaluwarsa",
    "status_inventory",
    "upload_foto",
    "upload_dokumen",
]

HEADER_ALIASES = {
    "nomor": "nomor",
    "kobar": "kobar",
    "reg": "reg",
    "reg.": "reg",
    "jenis_barang": "jenis_barang",
    "ukuran": "ukuran",
    "satuan": "satuan",
    "tgloleh": "tgloleh",
    "tgl_oleh": "tgloleh",
    "merk": "merk",
    "tipe": "tipe",
    "bahan": "bahan",
    "no_chasis_no_rangka": "rangka",
    "no_chasis/no_rangka": "rangka",
    "no_mesin_no_pabrik": "mesin",
    "no_mesin/no_pabrik": "mesin",
    "nomor_polisi": "nopol",
    "asal_oleh": "asal_oleh",
    "harga_rp": "harga",
    "harga_(rp)": "harga",
}

ALKES_OBJEK = {"70", "80", "90"}
ALKES_KEYWORDS = (
    "kedokteran",
    "laboratorium",
    "radiasi",
    "autoclave",
    "ultraviolet",
    "oksigen",
    "refractometer",
    "steril",
    "infra red",
    "infrared",
    "ultrasound",
    "x-ray",
    "rontgen",
    "defibrillator",
    "nebulizer",
    "ventilator",
    "tensimeter",
    "stetoskop",
)


def resolve_path(path: Path) -> Path:
    return path if path.is_absolute() else (ROOT / path)


def norm(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_header(value) -> str:
    text = norm(value).lower().replace(".", " ")
    text = re.sub(r"[^\w\s/()-]", "", text)
    text = re.sub(r"\s+", "_", text.strip()).replace("__", "_")
    return HEADER_ALIASES.get(text, text)


def digits(value) -> str:
    return re.sub(r"\D+", "", norm(value))


def parse_money(value) -> float:
    text = norm(value)
    if not text or text in {"-", "--"}:
        return 0.0
    text = text.replace(",", "")
    try:
        return float(text)
    except ValueError:
        return 0.0


def year_from(value) -> int | None:
    if isinstance(value, datetime):
        return value.year
    text = norm(value)
    if not text:
        return None
    match = re.search(r"(19|20)\d{2}", text)
    return int(match.group(0)) if match else None


def map_satuan_id(raw: str, id_unit: int, id_buah: int) -> int:
    code = norm(raw).upper()
    if code in {"BH", "BUAH", "B"}:
        return id_buah
    return id_unit


def jenis_barang_for(kobar: str, nama: str) -> str:
    objek = kobar[4:6] if len(kobar) >= 6 else ""
    nama_l = nama.lower()
    if objek in ALKES_OBJEK or any(key in nama_l for key in ALKES_KEYWORDS):
        return "ALKES"
    return "NON ALKES"


def blank_to_none(value: str | None) -> str | None:
    if value is None:
        return None
    text = norm(value)
    if text in {"", "-", "--"}:
        return None
    return text


def build_header_index(header_row) -> dict[str, int]:
    index: dict[str, int] = {}
    for i, cell in enumerate(header_row):
        key = normalize_header(cell)
        if key and key not in index:
            index[key] = i
    missing = [name for name in ("kobar", "jenis_barang") if name not in index]
    if missing:
        raise ValueError(
            "Header KIB tidak lengkap. Wajib: KOBAR, JENIS BARANG. "
            f"Hilang: {', '.join(missing)}"
        )
    return index


def pick_kib_sheet(workbook, sheet_arg: str | None):
    names = workbook.sheetnames
    if sheet_arg:
        if sheet_arg.isdigit():
            idx = int(sheet_arg)
            if idx < 0 or idx >= len(names):
                raise ValueError(f"Index sheet {idx} di luar jangkauan (0..{len(names) - 1})")
            return workbook[names[idx]]
        if sheet_arg not in names:
            raise ValueError(f"Sheet '{sheet_arg}' tidak ditemukan. Tersedia: {names}")
        return workbook[sheet_arg]

    for name in names:
        if "kib" in name.lower():
            return workbook[name]
    if len(names) >= 2:
        return workbook[names[1]]
    return workbook[names[0]]


def aggregate_inventory_rows(rows: list[list]) -> tuple[list[list], int]:
    """
    Gabungkan baris dengan kunci yang sama:
      id_data_barang (KOBAR) + merk + tahun_pembelian + harga_satuan (+ satuan + jenis)

    Tahun pembelian = tahun_produksi (histori KIB), fallback tahun_anggaran.
    Harga berbeda → tetap baris terpisah (meski merk & tahun sama).
    """
    if not rows:
        return [], 0

    # Index kolom sesuai HEADERS
    i_kobar = 0
    i_jenis = 5
    i_tahun_anggaran = 6
    i_qty = 7
    i_satuan = 8
    i_harga = 9
    i_merk = 10
    i_tipe = 11
    i_spek = 12
    i_tahun_produksi = 13
    i_penyedia = 14
    i_no_seri = 15

    grouped: dict[tuple, list] = {}
    order: list[tuple] = []

    for row in rows:
        kobar = digits(row[i_kobar]) if row[i_kobar] is not None else ""
        merk_key = norm(row[i_merk]).upper() if row[i_merk] is not None else ""
        tahun_beli = row[i_tahun_produksi] if row[i_tahun_produksi] not in (None, "") else row[i_tahun_anggaran]
        try:
            tahun_beli = int(tahun_beli) if tahun_beli is not None else 0
        except (TypeError, ValueError):
            tahun_beli = 0
        harga = round(float(row[i_harga] or 0), 2)
        satuan = int(row[i_satuan] or 0)
        jenis = norm(row[i_jenis]).upper()

        key = (kobar, merk_key, tahun_beli, harga, satuan, jenis)
        if key not in grouped:
            grouped[key] = list(row)
            order.append(key)
            continue

        base = grouped[key]
        try:
            base[i_qty] = int(base[i_qty] or 0) + int(row[i_qty] or 0)
        except (TypeError, ValueError):
            base[i_qty] = max(1, int(float(base[i_qty] or 1))) + max(1, int(float(row[i_qty] or 1)))

        # Setelah digabung: no_seri unik per unit tidak relevan
        base[i_no_seri] = None

        # Spesifikasi: pertahankan nama barang saja (buang nopol/rangka per-unit)
        spek = norm(base[i_spek])
        if " | " in spek:
            base[i_spek] = spek.split(" | ", 1)[0]

        # Tipe / penyedia: isi jika base kosong dan baris baru punya nilai
        if not base[i_tipe] and row[i_tipe]:
            base[i_tipe] = row[i_tipe]
        if not base[i_penyedia] and row[i_penyedia]:
            base[i_penyedia] = row[i_penyedia]

    merged = [grouped[k] for k in order]
    return merged, len(rows) - len(merged)


def convert_rows(
    source_rows: list,
    *,
    id_gudang: int,
    id_anggaran: int,
    id_sub_kegiatan: int | None,
    id_satuan_unit: int,
    id_satuan_buah: int,
    status: str,
    tahun_anggaran_default: int,
    aggregate: bool = True,
) -> tuple[list[list], int, int]:
    """
    Returns: (rows, skipped_source_rows, merged_away_count)
    """
    if not source_rows:
        raise ValueError("Sheet KIB kosong.")

    header_index = build_header_index(source_rows[0])
    out_rows: list[list] = []
    skipped = 0

    def cell(row, key: str, default=None):
        if key not in header_index:
            return default
        pos = header_index[key]
        return row[pos] if pos < len(row) else default

    for raw in source_rows[1:]:
        if not raw or all(v is None or str(v).strip() == "" for v in raw):
            skipped += 1
            continue

        kobar = digits(cell(raw, "kobar"))
        nama = norm(cell(raw, "jenis_barang"))
        if not kobar or not nama:
            skipped += 1
            continue

        tahun_histori = year_from(cell(raw, "tgloleh"))
        tahun_produksi = tahun_histori
        tahun_anggaran = tahun_histori or tahun_anggaran_default
        if tahun_anggaran < 2000 or tahun_anggaran > 2100:
            if tahun_produksi is None and 1900 <= tahun_anggaran <= 2100:
                tahun_produksi = tahun_anggaran
            tahun_anggaran = tahun_anggaran_default

        try:
            ukuran = cell(raw, "ukuran")
            qty = int(float(ukuran)) if ukuran not in (None, "") else 1
        except (TypeError, ValueError):
            qty = 1
        qty = max(1, qty)

        merk = blank_to_none(norm(cell(raw, "merk")))
        tipe = blank_to_none(norm(cell(raw, "tipe")))
        bahan = blank_to_none(norm(cell(raw, "bahan")))
        nopol = blank_to_none(norm(cell(raw, "nopol")))
        rangka = blank_to_none(norm(cell(raw, "rangka")))
        mesin = blank_to_none(norm(cell(raw, "mesin")))
        asal = blank_to_none(norm(cell(raw, "asal_oleh")))
        if asal in {"0", "1"}:
            asal = None

        # Spesifikasi per-unit (nopol/rangka) hanya relevan jika tidak digabung.
        # Saat aggregate=True, bagian unik akan dibuang di aggregate_inventory_rows.
        spec_parts = [nama]
        if bahan and bahan not in {"2"}:
            spec_parts.append(f"Bahan: {bahan}")
        if not aggregate:
            if nopol:
                spec_parts.append(f"Nopol: {nopol}")
            if rangka:
                spec_parts.append(f"Rangka: {rangka}")
            if mesin:
                spec_parts.append(f"Mesin/Pabrik: {mesin}")

        reg = blank_to_none(norm(cell(raw, "reg")))
        no_seri = None if aggregate else (reg or mesin)

        out_rows.append(
            [
                kobar,
                id_gudang,
                id_anggaran,
                id_sub_kegiatan,
                "ASET",
                jenis_barang_for(kobar, nama),
                tahun_anggaran,
                qty,
                map_satuan_id(norm(cell(raw, "satuan")), id_satuan_unit, id_satuan_buah),
                parse_money(cell(raw, "harga")),
                merk,
                tipe,
                " | ".join(spec_parts),
                tahun_produksi,
                asal,
                no_seri,
                None,
                None,
                status,
                None,
                None,
            ]
        )

    merged_away = 0
    if aggregate:
        out_rows, merged_away = aggregate_inventory_rows(out_rows)

    return out_rows, skipped, merged_away


def write_workbook(output: Path, rows: list[list], meta: dict) -> None:
    wb = Workbook()
    data_ws = wb.active
    data_ws.title = "data_inventory"
    data_ws.append(HEADERS)
    for row in rows:
        data_ws.append(row)

    tip = wb.create_sheet("petunjuk")
    tip.append(
        [
            f"Sumber: {meta['source']}",
            f"Sheet sumber: {meta['sheet']}",
            "jenis_inventory = ASET",
            "id_data_barang = kode KOBAR (importer resolve ke ID master)",
            f"id_gudang={meta['id_gudang']}, id_anggaran={meta['id_anggaran']}",
            "tahun_anggaran 2000–2100; tahun historis KIB di tahun_produksi",
            f"Total baris: {meta['rows']}. Dilewati sumber: {meta['skipped']}. Digabung: {meta.get('merged', 0)}.",
            f"ALKES={meta['alkes']}, NON ALKES={meta['non_alkes']}.",
            "Agregasi: KOBAR + merk + tahun pembelian + harga_satuan (harga beda → baris terpisah).",
            f"Dibuat: {meta['created_at']}",
            "Script: scripts/convert_kib_b_to_inventory_import.py",
        ]
    )

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Konversi KIB B → template import inventory ASET (data_inventory)."
    )
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE, help="Path file KIB B (.xlsx)")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Path file output import")
    parser.add_argument(
        "--sheet",
        default=None,
        help="Nama sheet atau index 0-based. Default: sheet berisi 'KIB' / sheet ke-2",
    )
    parser.add_argument("--id-gudang", type=int, default=1, help="ID gudang PUSAT ASET (default: 1)")
    parser.add_argument("--id-anggaran", type=int, default=1, help="ID sumber anggaran (default: 1)")
    parser.add_argument("--id-sub-kegiatan", type=int, default=None, help="ID sub kegiatan (opsional)")
    parser.add_argument("--id-satuan-unit", type=int, default=1, help="ID satuan Unit (default: 1)")
    parser.add_argument("--id-satuan-buah", type=int, default=3, help="ID satuan Buah/BH (default: 3)")
    parser.add_argument(
        "--status",
        default="AKTIF",
        choices=["DRAFT", "AKTIF", "DISTRIBUSI", "HABIS"],
        help="status_inventory (default: AKTIF)",
    )
    parser.add_argument(
        "--tahun-anggaran",
        type=int,
        default=datetime.now().year,
        help="Fallback tahun_anggaran jika TGLOLEH kosong/<2000 (default: tahun berjalan)",
    )
    parser.add_argument(
        "--no-aggregate",
        action="store_true",
        help="Jangan gabungkan baris (1 baris KIB = 1 baris import)",
    )
    parser.add_argument(
        "--aggregate-import",
        type=Path,
        default=None,
        help="Agregasi ulang file import inventory yang sudah ada (tanpa baca KIB B)",
    )
    return parser.parse_args()


def aggregate_existing_import(source: Path, output: Path) -> int:
    wb = load_workbook(source, data_only=True)
    sheet = wb[wb.sheetnames[0]]
    raw_rows = list(sheet.iter_rows(values_only=True))
    wb.close()
    if not raw_rows:
        print("File import kosong.", file=sys.stderr)
        return 1

    header = [norm(h).lower() for h in raw_rows[0]]
    expected = [h.lower() for h in HEADERS]
    if header[: len(expected)] != expected:
        # toleransi: pakai urutan HEADERS jika jumlah kolom cocok
        if len(raw_rows[0]) < len(HEADERS):
            print("Header file import tidak sesuai template data_inventory.", file=sys.stderr)
            return 1

    data = [list(r) for r in raw_rows[1:] if r and any(v is not None and str(v).strip() != "" for v in r)]
    # pastikan panjang kolom
    normalized = []
    for row in data:
        padded = list(row[: len(HEADERS)]) + [None] * max(0, len(HEADERS) - len(row))
        normalized.append(padded[: len(HEADERS)])

    before = len(normalized)
    merged, merged_away = aggregate_inventory_rows(normalized)
    alkes = sum(1 for row in merged if norm(row[5]).upper() == "ALKES")
    meta = {
        "source": str(source),
        "sheet": "data_inventory",
        "id_gudang": merged[0][1] if merged else "",
        "id_anggaran": merged[0][2] if merged else "",
        "rows": len(merged),
        "skipped": 0,
        "merged": merged_away,
        "alkes": alkes,
        "non_alkes": len(merged) - alkes,
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }
    write_workbook(output, merged, meta)
    print("Agregasi file import selesai.")
    print(f"  Sumber : {source}")
    print(f"  Output : {output}")
    print(f"  Sebelum: {before} → Sesudah: {len(merged)} (digabung {merged_away})")
    return 0


def main() -> int:
    args = parse_args()
    output = resolve_path(args.output)

    if args.aggregate_import:
        return aggregate_existing_import(resolve_path(args.aggregate_import), output)

    source = resolve_path(args.source)

    if not source.exists():
        print(f"File sumber tidak ditemukan: {source}", file=sys.stderr)
        return 1

    workbook = load_workbook(source, data_only=True)
    sheet = pick_kib_sheet(workbook, args.sheet)
    source_rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    rows, skipped, merged_away = convert_rows(
        source_rows,
        id_gudang=args.id_gudang,
        id_anggaran=args.id_anggaran,
        id_sub_kegiatan=args.id_sub_kegiatan,
        id_satuan_unit=args.id_satuan_unit,
        id_satuan_buah=args.id_satuan_buah,
        status=args.status,
        tahun_anggaran_default=args.tahun_anggaran,
        aggregate=not args.no_aggregate,
    )

    alkes = sum(1 for row in rows if row[5] == "ALKES")
    meta = {
        "source": str(source),
        "sheet": sheet.title,
        "id_gudang": args.id_gudang,
        "id_anggaran": args.id_anggaran,
        "rows": len(rows),
        "skipped": skipped,
        "merged": merged_away,
        "alkes": alkes,
        "non_alkes": len(rows) - alkes,
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }
    write_workbook(output, rows, meta)

    print("Konversi selesai.")
    print(f"  Sumber : {source} [{sheet.title}]")
    print(f"  Output : {output}")
    print(f"  Baris  : {len(rows)} (dilewati sumber {skipped}, digabung {merged_away})")
    print(f"  ALKES  : {alkes} | NON ALKES: {len(rows) - alkes}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
