from __future__ import annotations

from pathlib import Path
import argparse
import csv
import re
from openpyxl import load_workbook


def normalize_key(value: str) -> str:
    value = value.strip().lower().replace("-", "_").replace(" ", "_")
    return re.sub(r"[^a-z0-9_]", "", value)


def split_code(code: str) -> tuple[str, str, str, str, str, str]:
    digits = re.sub(r"\D+", "", code or "")
    digits = (digits[:12]).ljust(12, "0")
    return digits[0:1], digits[1:2], digits[2:4], digits[4:6], digits[6:8], digits[8:11]


def write_csv(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)


def main() -> int:
    parser = argparse.ArgumentParser(description="Convert BMD KIB B.xlsx to CSV import sheets")
    parser.add_argument("--source", default=r"C:\Users\pusda\Downloads\BMD KIB B.xlsx")
    parser.add_argument("--output-dir", default=r"E:\laragon\www\si-peni\storage\app\kemendagri_import_kib_b_csv")
    args = parser.parse_args()

    source = Path(args.source)
    output_dir = Path(args.output_dir)
    if not source.exists():
        raise FileNotFoundError(f"Source file not found: {source}")

    wb = load_workbook(source, read_only=True, data_only=True)
    if "DATA" not in wb.sheetnames:
        raise RuntimeError("Sheet DATA not found")
    ws = wb["DATA"]

    iterator = ws.iter_rows(values_only=True)
    header_row = next(iterator)
    headers = [normalize_key(str(h) if h is not None else "") for h in header_row]

    aset_rows, kode_rows, kategori_rows, jenis_rows, subjenis_rows, data_rows, perm_rows = [], [], [], [], [], [], []
    aset_seen, kode_seen, kategori_seen, jenis_seen, subjenis_seen, data_seen, perm_seen = set(), set(), set(), set(), set(), set(), set()

    for raw in iterator:
        mapped = {headers[i]: (raw[i] if i < len(raw) else None) for i in range(len(headers)) if headers[i]}
        kode = re.sub(r"\D+", "", str(mapped.get("kode_barang") or ""))
        nama_barang = str(mapped.get("nama_barang") or "").strip()
        if not kode or not nama_barang:
            continue

        objek = str(mapped.get("objek") or "").strip()
        rincian = str(mapped.get("rincian_objek") or "").strip()
        sub_rincian = str(mapped.get("sub_rincian_objek") or "").strip()
        deskripsi = str(mapped.get("deskripsi") or "").strip()
        id_tafsir = re.sub(r"\D+", "", str(mapped.get("id_tafsir") or ""))

        akun, kelompok, jenis, objek_kode, rincian_kode, sub_rincian_kode = split_code(kode)
        sub_sub = (id_tafsir or "0").zfill(3)

        nama_aset = "ASET TETAP" if akun == "1" else "ASET"
        kode_barang = f"{akun}.{kelompok}"
        kode_kategori = f"{kode_barang}.{jenis}"
        kode_jenis = f"{kode_kategori}.{objek_kode}"
        kode_subjenis = f"{kode_jenis}.{rincian_kode}"
        kode_data = kode

        if nama_aset not in aset_seen:
            aset_rows.append([nama_aset]); aset_seen.add(nama_aset)
        if kode_barang not in kode_seen:
            kode_rows.append([kode_barang, objek or f"OBJEK {kode_barang}", nama_aset]); kode_seen.add(kode_barang)
        if kode_kategori not in kategori_seen:
            kategori_rows.append([kode_kategori, rincian or f"RINCIAN {kode_kategori}", kode_barang]); kategori_seen.add(kode_kategori)
        if kode_jenis not in jenis_seen:
            jenis_rows.append([kode_jenis, sub_rincian or f"SUB RINCIAN {kode_jenis}", kode_kategori]); jenis_seen.add(kode_jenis)
        if kode_subjenis not in subjenis_seen:
            subjenis_rows.append([kode_subjenis, sub_rincian or f"SUBJENIS {kode_subjenis}", kode_jenis]); subjenis_seen.add(kode_subjenis)
        if kode_data not in data_seen:
            data_rows.append([kode_data, nama_barang, deskripsi or "", kode_subjenis, "Unit", ""]); data_seen.add(kode_data)
        if kode_data not in perm_seen:
            perm_rows.append([kode_data, akun, kelompok, jenis, objek_kode, rincian_kode, sub_rincian_kode, sub_sub, "REVIEW", "Dikonversi otomatis dari BMD KIB B"]); perm_seen.add(kode_data)

    write_csv(output_dir / "aset.csv", ["nama_aset"], aset_rows)
    write_csv(output_dir / "kode_barang.csv", ["kode_barang", "nama_kode_barang", "nama_aset"], kode_rows)
    write_csv(output_dir / "kategori_barang.csv", ["kode_kategori_barang", "nama_kategori_barang", "kode_barang"], kategori_rows)
    write_csv(output_dir / "jenis_barang.csv", ["kode_jenis_barang", "nama_jenis_barang", "kode_kategori_barang"], jenis_rows)
    write_csv(output_dir / "subjenis_barang.csv", ["kode_subjenis_barang", "nama_subjenis_barang", "kode_jenis_barang"], subjenis_rows)
    write_csv(output_dir / "data_barang.csv", ["kode_data_barang", "nama_barang", "deskripsi", "kode_subjenis_barang", "nama_satuan", "foto_barang"], data_rows)
    write_csv(output_dir / "permendagri_108.csv", ["kode_data_barang", "kode_akun", "kode_kelompok", "kode_jenis_108", "kode_objek", "kode_rincian_objek", "kode_sub_rincian_objek", "kode_sub_sub_rincian_objek", "status_validasi", "catatan"], perm_rows)

    print(f"Done: {output_dir}")
    print(f"data_barang rows: {len(data_rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

