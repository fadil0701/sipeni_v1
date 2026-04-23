from __future__ import annotations

from pathlib import Path
import argparse
import re
from openpyxl import Workbook, load_workbook


def normalize_key(value: str) -> str:
    value = value.strip().lower().replace("-", "_").replace(" ", "_")
    return re.sub(r"[^a-z0-9_]", "", value)


def split_code(code: str) -> tuple[str, str, str, str, str, str]:
    digits = re.sub(r"\D+", "", code or "")
    digits = (digits[:12]).ljust(12, "0")
    return (
        digits[0:1],   # akun
        digits[1:2],   # kelompok
        digits[2:4],   # jenis
        digits[4:6],   # objek
        digits[6:8],   # rincian
        digits[8:11],  # sub rincian
    )


def main() -> int:
    parser = argparse.ArgumentParser(description="Convert BMD KIB B.xlsx to system import template")
    parser.add_argument("--source", default=r"C:\Users\pusda\Downloads\BMD KIB B.xlsx")
    parser.add_argument("--output", default=r"E:\laragon\www\si-peni\storage\app\kemendagri_import_kib_b.xlsx")
    args = parser.parse_args()

    source = Path(args.source)
    output = Path(args.output)

    if not source.exists():
        raise FileNotFoundError(f"Source file not found: {source}")

    wb = load_workbook(source, read_only=True, data_only=True)
    if "DATA" not in wb.sheetnames:
        raise RuntimeError("Sheet 'DATA' not found in source workbook")

    ws = wb["DATA"]
    rows = ws.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        raise RuntimeError("DATA sheet is empty")

    headers = [normalize_key(str(h) if h is not None else "") for h in header_row]

    aset_rows: list[list[str]] = []
    kode_rows: list[list[str]] = []
    kategori_rows: list[list[str]] = []
    jenis_rows: list[list[str]] = []
    subjenis_rows: list[list[str]] = []
    data_rows: list[list[str | None]] = []
    permendagri_rows: list[list[str]] = []

    aset_seen: set[str] = set()
    kode_seen: set[str] = set()
    kategori_seen: set[str] = set()
    jenis_seen: set[str] = set()
    subjenis_seen: set[str] = set()
    data_seen: set[str] = set()
    perm_seen: set[str] = set()

    for raw in rows:
        mapped = {headers[i]: (raw[i] if i < len(raw) else None) for i in range(len(headers)) if headers[i]}
        kode = re.sub(r"\D+", "", str(mapped.get("kode_barang") or ""))
        nama_barang = str(mapped.get("nama_barang") or "").strip()
        objek = str(mapped.get("objek") or "").strip()
        rincian = str(mapped.get("rincian_objek") or "").strip()
        sub_rincian = str(mapped.get("sub_rincian_objek") or "").strip()
        deskripsi = str(mapped.get("deskripsi") or "").strip()
        id_tafsir = re.sub(r"\D+", "", str(mapped.get("id_tafsir") or ""))

        if not kode or not nama_barang:
            continue

        akun, kelompok, jenis, objek_kode, rincian_kode, sub_rincian_kode = split_code(kode)
        sub_sub_rincian_kode = (id_tafsir or "0").zfill(3)

        nama_aset = "ASET TETAP" if akun == "1" else "ASET"
        kode_barang = f"{akun}.{kelompok}"
        kode_kategori = f"{kode_barang}.{jenis}"
        kode_jenis = f"{kode_kategori}.{objek_kode}"
        kode_subjenis = f"{kode_jenis}.{rincian_kode}"
        kode_data_barang = kode

        if nama_aset not in aset_seen:
            aset_rows.append([nama_aset])
            aset_seen.add(nama_aset)

        if kode_barang not in kode_seen:
            kode_rows.append([kode_barang, objek or f"OBJEK {kode_barang}", nama_aset])
            kode_seen.add(kode_barang)

        if kode_kategori not in kategori_seen:
            kategori_rows.append([kode_kategori, rincian or f"RINCIAN {kode_kategori}", kode_barang])
            kategori_seen.add(kode_kategori)

        if kode_jenis not in jenis_seen:
            jenis_rows.append([kode_jenis, sub_rincian or f"SUB RINCIAN {kode_jenis}", kode_kategori])
            jenis_seen.add(kode_jenis)

        if kode_subjenis not in subjenis_seen:
            subjenis_rows.append([kode_subjenis, sub_rincian or f"SUBJENIS {kode_subjenis}", kode_jenis])
            subjenis_seen.add(kode_subjenis)

        if kode_data_barang not in data_seen:
            data_rows.append([kode_data_barang, nama_barang, deskripsi or None, kode_subjenis, "Unit", None])
            data_seen.add(kode_data_barang)

        if kode_data_barang not in perm_seen:
            permendagri_rows.append([
                kode_data_barang,
                akun,
                kelompok,
                jenis,
                objek_kode,
                rincian_kode,
                sub_rincian_kode,
                sub_sub_rincian_kode,
                "REVIEW",
                "Dikonversi otomatis dari BMD KIB B",
            ])
            perm_seen.add(kode_data_barang)

    out = Workbook()

    def fill_sheet(ws_out, title: str, headers_out: list[str], body: list[list[str | None]]) -> None:
        ws_out.title = title
        ws_out.append(headers_out)
        for r in body:
            ws_out.append(r)

    fill_sheet(out.active, "aset", ["nama_aset"], aset_rows)
    fill_sheet(out.create_sheet(), "kode_barang", ["kode_barang", "nama_kode_barang", "nama_aset"], kode_rows)
    fill_sheet(out.create_sheet(), "kategori_barang", ["kode_kategori_barang", "nama_kategori_barang", "kode_barang"], kategori_rows)
    fill_sheet(out.create_sheet(), "jenis_barang", ["kode_jenis_barang", "nama_jenis_barang", "kode_kategori_barang"], jenis_rows)
    fill_sheet(out.create_sheet(), "subjenis_barang", ["kode_subjenis_barang", "nama_subjenis_barang", "kode_jenis_barang"], subjenis_rows)
    fill_sheet(out.create_sheet(), "data_barang", ["kode_data_barang", "nama_barang", "deskripsi", "kode_subjenis_barang", "nama_satuan", "foto_barang"], data_rows)
    fill_sheet(out.create_sheet(), "permendagri_108", [
        "kode_data_barang",
        "kode_akun",
        "kode_kelompok",
        "kode_jenis_108",
        "kode_objek",
        "kode_rincian_objek",
        "kode_sub_rincian_objek",
        "kode_sub_sub_rincian_objek",
        "status_validasi",
        "catatan",
    ], permendagri_rows)

    output.parent.mkdir(parents=True, exist_ok=True)
    out.save(output)

    print(f"Done: {output}")
    print(f"Rows data_barang: {len(data_rows)}")
    print(f"Rows permendagri_108: {len(permendagri_rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

