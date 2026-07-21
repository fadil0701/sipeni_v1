"""
Lengkapi hierarki Aset Lancar (Persediaan + Farmasi) pada file import Kemendagri
sampai level data_barang, dengan item ATK / Cetakan / ART dari Stock Opname.

Usage:
  py scripts/enrich_kemendagri_persediaan_farmasi.py
  py scripts/enrich_kemendagri_persediaan_farmasi.py --target "database/seeders/data/....xlsx"
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_TARGET = ROOT / (
    "database/seeders/data/"
    "kemendagri_import_sheet6_objek_filtered_v2.backup_before_kib_bmd_20260720_085614.xlsx"
)
DEFAULT_STOCK = ROOT / "database/seeders/data/Stock Opname ATK CETAKAN ART 2026 (5).xlsx"
STOCK_SHEET = "SO ATK,CETAKAN, ART - T.A 2026"

# Hierarki Permendagri/BMD (diselaraskan dengan bmd_aset_lancar_lengkap.xlsx)
# + perluasan Cetakan & Alat Rumah Tangga (belum ada di BMD ringkas).
KATEGORI = [
    ("1.1.01", "Barang Habis Pakai", "1.1"),
    ("1.1.02", "Bahan Material", "1.1"),
    ("1.1.03", "Suku Cadang", "1.1"),
    ("1.1.04", "Barang Diserahkan", "1.1"),
    ("1.1.05", "Barang Proses", "1.1"),
    ("1.1.06", "Lainnya", "1.1"),
]

JENIS = [
    ("1.1.01.01", "Persediaan Kantor", "1.1.01"),
    ("1.1.01.02", "Persediaan Farmasi", "1.1.01"),
    ("1.1.02.01", "Bahan Material", "1.1.02"),
    ("1.1.03.01", "Suku Cadang", "1.1.03"),
    ("1.1.04.01", "Barang Diserahkan", "1.1.04"),
    ("1.1.05.01", "Barang Proses", "1.1.05"),
    ("1.1.06.01", "Lainnya", "1.1.06"),
]

SUBJENIS = [
    ("1.1.01.01.01", "Alat Tulis Kantor", "1.1.01.01"),
    ("1.1.01.01.02", "Kertas", "1.1.01.01"),
    ("1.1.01.01.03", "Tinta/Toner", "1.1.01.01"),
    ("1.1.01.01.04", "Bahan Cetakan", "1.1.01.01"),
    ("1.1.01.01.05", "Alat Rumah Tangga", "1.1.01.01"),
    ("1.1.01.02.01", "Obat-obatan Umum", "1.1.01.02"),
    ("1.1.01.02.02", "Obat Resep", "1.1.01.02"),
    ("1.1.01.02.03", "Obat Generik", "1.1.01.02"),
    ("1.1.01.02.04", "Vaksin", "1.1.01.02"),
    ("1.1.01.02.05", "BMHP", "1.1.01.02"),
    ("1.1.02.01.01", "Bahan Bangunan", "1.1.02.01"),
    ("1.1.02.01.02", "Bahan Kayu", "1.1.02.01"),
    ("1.1.03.01.01", "Suku Cadang Kendaraan", "1.1.03.01"),
    ("1.1.03.01.02", "Suku Cadang Elektronik", "1.1.03.01"),
    ("1.1.04.01.01", "Bantuan Sosial", "1.1.04.01"),
    ("1.1.04.01.02", "Hibah", "1.1.04.01"),
    ("1.1.05.01.01", "Barang Dalam Proses", "1.1.05.01"),
    ("1.1.06.01.01", "Persediaan Lainnya", "1.1.06.01"),
]

# Placeholder data_barang untuk subjenis tanpa daftar item operasional.
PLACEHOLDER_DATA = [
    ("110102010000", "Obat-obatan Umum", "1.1.01.02.01", "Unit"),
    ("110102020000", "Obat Resep", "1.1.01.02.02", "Unit"),
    ("110102030000", "Obat Generik", "1.1.01.02.03", "Unit"),
    ("110102040000", "Vaksin", "1.1.01.02.04", "Unit"),
    ("110102050000", "BMHP", "1.1.01.02.05", "Unit"),
    ("110201010000", "Bahan Bangunan", "1.1.02.01.01", "Unit"),
    ("110201020000", "Bahan Kayu", "1.1.02.01.02", "Unit"),
    ("110301010000", "Suku Cadang Kendaraan", "1.1.03.01.01", "Unit"),
    ("110301020000", "Suku Cadang Elektronik", "1.1.03.01.02", "Unit"),
    ("110401010000", "Bantuan Sosial", "1.1.04.01.01", "Unit"),
    ("110401020000", "Hibah", "1.1.04.01.02", "Unit"),
    ("110501010000", "Barang Dalam Proses", "1.1.05.01.01", "Unit"),
    ("110601010000", "Persediaan Lainnya", "1.1.06.01.01", "Unit"),
]

SECTION_ALIASES = {
    "ALAT TULIS KANTOR": "ATK",
    "CETAKAN UMUM": "CETAKAN",
    "CETAKAN KHUSUS": "CETAKAN",
    "CETAKAN": "CETAKAN",
    "ALAT RUMAH TANGGA (ART)": "ART",
    "ALAT RUMAH TANGGA": "ART",
    "ART": "ART",
}

JUNK_NAMA = {
    "ATK",
    "ART",
    "CETAKAN",
    "CETAKAN UMUM",
    "CETAKAN KHUSUS",
    "ALAT TULIS KANTOR",
    "ALAT RUMAH TANGGA",
    "ALAT RUMAH TANGGA (ART)",
}


def norm_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


# Canonical names dari MasterSatuanSeeder (case-sensitive match di DB).
SATUAN_MAP: dict[str, str] = {
    "UNIT": "Unit",
    "PCS": "PCS",
    "PC": "PCS",
    "BUAH": "Buah",
    "BH": "Buah",
    "PASANG": "Pasang",
    "SET": "Set",
    "LUSIN": "Lusin",
    "LSN": "Lusin",
    "KODI": "Kodi",
    "GROSS": "Gross",
    "RIM": "Rim",
    "LEM": "Lembar",
    "LEMBAR": "Lembar",
    "LBR": "Lembar",
    "LBR.": "Lembar",
    "ROLL": "Roll",
    "ROL": "Roll",
    "IKAT": "Ikat",
    "BATANG": "Batang",
    "POTONG": "Potong",
    "BIJI": "Biji",
    "HELAI": "Helai",
    "BOX": "Box",
    "BOXS": "Box",
    "KOTAK": "Box",
    "DUS": "Dus",
    "KARTON": "Karton",
    "PACK": "Pak",
    "PAK": "Pak",
    "PACKET": "Pak",
    "BUNDLE": "Bundle",
    "PALLET": "Pallet",
    "KARUNG": "Karung",
    "KALENG": "Kaleng",
    "KLENG": "Kaleng",  # typo di Stock Opname
    "TABUNG": "Tabung",
    "JERIGEN": "Jerigen",
    "DRUM": "Drum",
    "BAG": "Bag",
    "STRIP": "Strip",
    "BLISTER": "Blister",
    "SACHET": "Sachet",
    "POUCH": "Pouch",
    "TUBE": "Tube",
    "BOTOL": "Botol",
    "BTL": "Botol",
    "BTL.": "Botol",
    "AMPUL": "Ampul",
    "VIAL": "Vial",
    "TABLET": "Tablet",
    "KAPLET": "Kaplet",
    "KAPSUL": "Kapsul",
    "KG": "kg",
    "G": "g",
    "MG": "mg",
    "L": "L",
    "ML": "mL",
    "CC": "cc",
    "BUKU": "Book",
    "BOOK": "Book",
    "PAD": "Pad",
    "DOSIS": "Dosis",
    "KIT": "Kit",
    "TEST": "Test",
    # Gabungan yang sering muncul di SO
    "RIM/BUKU": "Rim",
    "RIM / BUKU": "Rim",
}


def normalize_satuan(value: str) -> str:
    """Samakan satuan Stock Opname ke nama di MasterSatuanSeeder."""
    raw = re.sub(r"\s+", " ", str(value or "").strip())
    if not raw or raw.upper() in {"SATUAN", "-", "N/A", "NA"}:
        return "Unit"

    key = raw.upper()
    if key in SATUAN_MAP:
        return SATUAN_MAP[key]

    # Coba tanpa tanda baca di ujung
    key2 = key.strip(" .")
    if key2 in SATUAN_MAP:
        return SATUAN_MAP[key2]

    # Fallback: Title Case untuk teks panjang, biarkan singkat apa adanya
    if len(raw) <= 3:
        return SATUAN_MAP.get(raw.upper(), raw.upper() if raw.isupper() else raw)
    return raw.title()


def classify_atk_item(nama: str) -> str:
    u = nama.upper()
    if any(
        k in u
        for k in (
            "TINTA ",
            "TINTA/",
            "TONER",
            "RIBBON PRINTER",
        )
    ) or u.startswith("TINTA") or u.startswith("TONER"):
        return "1.1.01.01.03"
    if any(
        k in u
        for k in (
            "KERTAS HVS",
            "KERTAS KARTON",
            "KERTAS POS IT",
            "KERTAS PRONTO",
            "COVER PLASTIK",
            "LABEL BARCODE THERMAL",
        )
    ):
        return "1.1.01.01.02"
    return "1.1.01.01.01"


def subjenis_to_segments(kode_subjenis: str) -> tuple[str, str, str, str, str]:
    # 1.1.01.01.04 -> akun=1, kelompok=1, jenis=01, objek=01, rincian=04
    parts = kode_subjenis.split(".")
    if len(parts) != 5:
        raise ValueError(f"Kode subjenis tidak valid: {kode_subjenis}")
    return parts[0], parts[1], parts[2], parts[3], parts[4]


def make_kode_data(kode_subjenis: str, seq: int) -> str:
    akun, kelompok, jenis, objek, rincian = subjenis_to_segments(kode_subjenis)
    return f"{akun}{kelompok}{jenis}{objek}{rincian}{seq:04d}"


def make_permendagri_row(kode_data: str, kode_subjenis: str, seq: int, catatan: str) -> tuple:
    akun, kelompok, jenis, objek, rincian = subjenis_to_segments(kode_subjenis)
    return (
        kode_data,
        akun,
        kelompok,
        jenis,
        objek,
        rincian,
        f"{seq:03d}",
        f"{seq:03d}",
        "REVIEW",
        catatan,
    )


def sheet_key_set(ws) -> set[str]:
    keys = set()
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] is not None and str(row[0]).strip() != "":
            keys.add(str(row[0]).strip())
    return keys


def upsert_rows(ws, key_col_idx: int, rows: list[tuple], replace_prefix: str | None = None) -> int:
    """Append rows whose key belum ada. Jika replace_prefix, hapus dulu baris dengan key berawalan itu."""
    if replace_prefix:
        to_delete = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=True), start=2):
            val = str(row[0] or "").strip()
            if val.startswith(replace_prefix):
                to_delete.append(idx)
        for idx in reversed(to_delete):
            ws.delete_rows(idx, 1)

    existing = sheet_key_set(ws)
    added = 0
    for row in rows:
        key = str(row[0]).strip()
        if key in existing:
            continue
        ws.append(list(row))
        existing.add(key)
        added += 1
    return added


def parse_stock_items(path: Path) -> dict[str, list[tuple[str, str]]]:
    wb = load_workbook(path, data_only=True)
    if STOCK_SHEET not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{STOCK_SHEET}' tidak ditemukan di {path}")
    ws = wb[STOCK_SHEET]

    current: str | None = None
    last_nama: str | None = None
    buckets: dict[str, list[tuple[str, str]]] = {"ATK": [], "CETAKAN": [], "ART": []}

    for row in ws.iter_rows(min_row=1, max_col=4, values_only=True):
        no, _kode, nama, satuan = row
        nama_s = re.sub(r"\s+", " ", str(nama).strip()) if nama else ""
        satuan_s = str(satuan).strip() if satuan else ""
        no_s = str(no).strip() if no is not None else ""

        if nama_s:
            upper = nama_s.upper()
            if upper in SECTION_ALIASES or any(upper.startswith(k) for k in SECTION_ALIASES):
                # Section header: Roman / roman-like no, or plain category name
                if (
                    re.fullmatch(r"[IVX]+", no_s)
                    or no_s == ""
                    or no_s == "."
                    or upper in SECTION_ALIASES
                ):
                    for alias, cat in SECTION_ALIASES.items():
                        if upper == alias or upper.startswith(alias):
                            current = cat
                            last_nama = None
                            break
                    continue

        if current is None:
            continue

        if nama_s.upper().startswith("JUMLAH TOTAL"):
            current = None
            last_nama = None
            continue

        if not satuan_s:
            continue

        # Baris lanjutan harga (nama kosong) → skip; butuh nama
        if not nama_s:
            continue

        if nama_s.upper() in JUNK_NAMA:
            continue
        if re.fullmatch(r"[\d.,]+", nama_s):
            continue

        # Hindari duplikat nama (beda harga/tahun) dalam kategori yang sama
        buckets[current].append((nama_s, normalize_satuan(satuan_s)))
        last_nama = nama_s

    wb.close()

    # Unique by upper nama, keep first satuan
    unique: dict[str, list[tuple[str, str]]] = {}
    for cat, items in buckets.items():
        seen: set[str] = set()
        out: list[tuple[str, str]] = []
        for nama, satuan in items:
            key = nama.upper()
            if key in seen:
                continue
            seen.add(key)
            out.append((nama, satuan))
        unique[cat] = out
    return unique


def satuan_for_item(nama: str, satuan: str, kode_subjenis: str) -> str:
    """Override satuan agar selaras kebiasaan stok (tinta/toner → PCS)."""
    u = nama.upper()
    # Cartridge tinta/toner printer: di SO sering "UNIT", di master stok pakai PCS
    if kode_subjenis == "1.1.01.01.03":
        if any(k in u for k in ("TINTA HP", "TINTA EPSON", "TONER", "RIBBON")):
            if satuan in {"Unit", "Buah"}:
                return "PCS"
        # Tinta stempel botol tetap Botol
    return satuan


def build_item_rows(stock: dict[str, list[tuple[str, str]]]) -> tuple[list[tuple], list[tuple]]:
    data_rows: list[tuple] = []
    perm_rows: list[tuple] = []
    counters: dict[str, int] = {s[0]: 0 for s in SUBJENIS}

    def add_item(nama: str, satuan: str, kode_subjenis: str, catatan: str) -> None:
        satuan = satuan_for_item(nama, satuan, kode_subjenis)
        counters[kode_subjenis] = counters.get(kode_subjenis, 0) + 1
        seq = counters[kode_subjenis]
        kode = make_kode_data(kode_subjenis, seq)
        data_rows.append((kode, nama, None, kode_subjenis, satuan, None))
        perm_rows.append(make_permendagri_row(kode, kode_subjenis, seq, catatan))

    for nama, satuan in stock.get("ATK", []):
        add_item(nama, satuan, classify_atk_item(nama), "Stock Opname ATK 2026")

    for nama, satuan in stock.get("CETAKAN", []):
        add_item(nama, satuan, "1.1.01.01.04", "Stock Opname Cetakan 2026")

    for nama, satuan in stock.get("ART", []):
        add_item(nama, satuan, "1.1.01.01.05", "Stock Opname ART 2026")

    for kode, nama, sub, satuan in PLACEHOLDER_DATA:
        data_rows.append((kode, nama, None, sub, satuan, None))
        # Placeholder memakai seq 0
        perm_rows.append(make_permendagri_row(kode, sub, 0, "Hierarki Persediaan/Farmasi (placeholder)"))

    return data_rows, perm_rows


def ensure_headers(ws, expected: list[str]) -> None:
    headers = [norm_header(c.value) for c in ws[1]]
    expected_n = [norm_header(h) for h in expected]
    if headers[: len(expected_n)] != expected_n:
        raise RuntimeError(f"Header sheet '{ws.title}' tidak sesuai. Ada={headers} Harap={expected_n}")


def enrich(target: Path, stock_path: Path) -> dict[str, int]:
    stock = parse_stock_items(stock_path)
    data_rows, perm_rows = build_item_rows(stock)

    wb = load_workbook(target)

    required = [
        "aset",
        "kode_barang",
        "kategori_barang",
        "jenis_barang",
        "subjenis_barang",
        "data_barang",
        "permendagri_108",
    ]
    for name in required:
        if name not in wb.sheetnames:
            raise RuntimeError(f"Sheet '{name}' tidak ada di {target}")

    # aset
    ensure_headers(wb["aset"], ["nama_aset"])
    aset_existing = {str(r[0]).strip().upper() for r in wb["aset"].iter_rows(min_row=2, max_col=1, values_only=True) if r[0]}
    if "ASET LANCAR" not in aset_existing:
        wb["aset"].append(["ASET LANCAR"])

    # kode_barang
    ensure_headers(wb["kode_barang"], ["kode_barang", "nama_kode_barang", "nama_aset"])
    upsert_rows(wb["kode_barang"], 0, [("1.1", "Persediaan", "ASET LANCAR")])

    # kategori / jenis / subjenis — replace existing 1.1.* then insert fresh
    ensure_headers(wb["kategori_barang"], ["kode_kategori_barang", "nama_kategori_barang", "kode_barang"])
    ensure_headers(wb["jenis_barang"], ["kode_jenis_barang", "nama_jenis_barang", "kode_kategori_barang"])
    ensure_headers(wb["subjenis_barang"], ["kode_subjenis_barang", "nama_subjenis_barang", "kode_jenis_barang"])
    n_kat = upsert_rows(wb["kategori_barang"], 0, KATEGORI, replace_prefix="1.1.")
    n_jen = upsert_rows(wb["jenis_barang"], 0, JENIS, replace_prefix="1.1.")
    n_sub = upsert_rows(wb["subjenis_barang"], 0, SUBJENIS, replace_prefix="1.1.")

    # data_barang & permendagri_108: hapus semua kode 11* (aset lancar lama) lalu append baru
    ensure_headers(
        wb["data_barang"],
        ["kode_data_barang", "nama_barang", "deskripsi", "kode_subjenis_barang", "nama_satuan", "foto_barang"],
    )
    ensure_headers(
        wb["permendagri_108"],
        [
            "kode_data_barang",
            "kode_akun",
            "kode_kelompok",
            "kode_jenis_108",
            "kode_objek",
            "kode_rincian_objek",
            "kode_sub_rincian_objek",
            "kode_sub_sub_rincian_objek",
            "status_validasi",
            "catatan",
        ],
    )

    n_data = upsert_rows(wb["data_barang"], 0, data_rows, replace_prefix="11")
    n_perm = upsert_rows(wb["permendagri_108"], 0, perm_rows, replace_prefix="11")

    wb.save(target)
    wb.close()

    return {
        "atk": len(stock.get("ATK", [])),
        "cetakan": len(stock.get("CETAKAN", [])),
        "art": len(stock.get("ART", [])),
        "kategori_added": n_kat,
        "jenis_added": n_jen,
        "subjenis_added": n_sub,
        "data_barang_added": n_data,
        "permendagri_added": n_perm,
        "data_barang_total_new": len(data_rows),
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--target", type=Path, default=DEFAULT_TARGET)
    parser.add_argument("--stock", type=Path, default=DEFAULT_STOCK)
    args = parser.parse_args()

    if not args.target.exists():
        raise SystemExit(f"Target tidak ditemukan: {args.target}")
    if not args.stock.exists():
        raise SystemExit(f"Stock Opname tidak ditemukan: {args.stock}")

    stats = enrich(args.target.resolve(), args.stock.resolve())
    print("Selesai memperkaya file Kemendagri.")
    print(f"  Target : {args.target}")
    print(f"  ATK    : {stats['atk']} item")
    print(f"  Cetakan: {stats['cetakan']} item")
    print(f"  ART    : {stats['art']} item")
    print(f"  Hierarki ditambah kategori/jenis/subjenis: {stats['kategori_added']}/{stats['jenis_added']}/{stats['subjenis_added']}")
    print(f"  data_barang baru: {stats['data_barang_added']} (total baris persediaan/farmasi: {stats['data_barang_total_new']})")
    print(f"  permendagri_108 baru: {stats['permendagri_added']}")


if __name__ == "__main__":
    main()
